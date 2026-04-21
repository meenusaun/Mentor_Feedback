import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from collections import defaultdict
import anthropic
from openai import OpenAI
import os
import json
import re

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Resources Network – Mentor Feedback Tracker",
    page_icon="📋",
    layout="wide"
)

# ─────────────────────────────────────────────
# HEADER  (same style as Mentor Search)
# ─────────────────────────────────────────────
col1, col2, col3 = st.columns([1, 2, 1])
with col1:
    if os.path.exists("DP_BG1.png"):
        st.image("DP_BG1.png", width=150)
with col2:
    st.markdown(
        "<h2 style='text-align: center;'>📋 Resources Network – Mentor Feedback Tracker</h2>",
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────
# CLIENTS
# ─────────────────────────────────────────────
@st.cache_resource
def get_clients():
    openai_key = st.secrets.get("OPENAI_API_KEY", os.environ.get("OPENAI_API_KEY", ""))
    anthropic_key = st.secrets.get("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))
    oc = OpenAI(api_key=openai_key) if openai_key else None
    ac = anthropic.Anthropic(api_key=anthropic_key) if anthropic_key else None
    return oc, ac

openai_client, anthropic_client = get_clients()

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
st.sidebar.title("⚙️ Settings")

ai_model = st.sidebar.radio(
    "AI Model for Insights:",
    options=["GPT-4o Mini (OpenAI)", "Claude (Anthropic)"],
    index=1,
)
st.sidebar.markdown("---")

st.sidebar.subheader("📂 Upload Tracker Excel")
st.sidebar.caption("Upload the Accelerate Mentor Connect Tracker (.xlsx)")
uploaded_file = st.sidebar.file_uploader(
    "Tracker File", type=["xlsx"], key="tracker_upload"
)

st.sidebar.markdown("---")
if st.sidebar.button("🗑️ Clear / Reset"):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()

# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────
RATING_MAP = {
    "Extremely useful": "Good",
    "Very useful":      "Good",
    "Moderately useful":"Moderate",
    "Slightly useful":  "Poor",
    "Not useful":       "Poor",
}

def classify(u):
    if u in ("Extremely useful", "Very useful"):   return "Good"
    if u == "Moderately useful":                    return "Moderate"
    if u in ("Slightly useful", "Not useful"):      return "Poor"
    return None

def load_tracker(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # ── Ventures lookup (hub / program) ──────────────────────────────
    venture_meta = {}
    if "Ventures" in wb.sheetnames:
        ws_v = wb["Ventures"]
        for r in ws_v.iter_rows(values_only=True):
            if r[0] and str(r[0]).strip():
                venture_meta[str(r[0]).strip()] = {
                    "program": str(r[1]).strip() if r[1] else "",
                    "hub":     str(r[3]).strip() if r[3] else "",
                }

    # ── Mentor Connects: completed meetings per mentor ────────────────
    completed_map = defaultdict(int)
    if "Mentor Connects" in wb.sheetnames:
        ws_mc = wb["Mentor Connects"]
        for i, r in enumerate(ws_mc.iter_rows(values_only=True)):
            if i == 0:
                continue
            mentor = r[4]
            status = r[5]
            if mentor and str(status).strip() == "Connected":
                completed_map[str(mentor).strip()] += 1

    # ── Feedback from Founders ────────────────────────────────────────
    rows = []
    if "Feedback from Founders" not in wb.sheetnames:
        return pd.DataFrame(), completed_map, venture_meta

    ws_fb = wb["Feedback from Founders"]
    for i, r in enumerate(ws_fb.iter_rows(values_only=True)):
        if i == 0:
            continue

        response_status = r[22]
        connected_by_us = r[23]

        # Skip only explicit duplicates
        if (str(response_status).strip() == "Duplicate" or
                str(connected_by_us).strip() == "Duplicate"):
            continue

        mentor   = r[15]
        venture  = r[11]
        rating   = r[16]
        cat      = classify(str(rating).strip() if rating else "")

        if not mentor or not venture or not cat:
            continue

        venture_str = str(venture).strip()
        meta = venture_meta.get(venture_str, {})

        rows.append({
            "mentor":       str(mentor).strip(),
            "venture":      venture_str,
            "rating_raw":   str(rating).strip() if rating else "",
            "rating":       cat,
            "by_rn":        str(connected_by_us).strip() == "Yes",
            "feedback":     str(r[19]).strip() if r[19] and str(r[19]).strip() not in ["None",""] else "",
            "rn_remarks":   str(r[20]).strip() if r[20] and str(r[20]).strip() not in ["None",""] else "",
            "action_items": str(r[17]).strip() if r[17] and str(r[17]).strip() not in ["None",""] else "",
            "meet_again":   str(r[18]).strip() if r[18] else "",
            "program":      meta.get("program", ""),
            "hub":          meta.get("hub", ""),
            "founder_name": str(r[10]).strip() if r[10] else "",
        })

    df = pd.DataFrame(rows)
    return df, completed_map, venture_meta


# ─────────────────────────────────────────────
# AI CALL HELPER
# ─────────────────────────────────────────────
def call_ai(prompt, max_tokens=1500):
    if ai_model == "GPT-4o Mini (OpenAI)" and openai_client:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )
        return resp.choices[0].message.content
    elif anthropic_client:
        resp = anthropic_client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=max_tokens,
            temperature=0,
            system=(
                "You are an intelligent program operations assistant for NEN (National "
                "Entrepreneurship Network). You help the Resources Network team analyse "
                "mentor-founder session feedback, identify patterns, flag concerns, and "
                "surface actionable insights. Be concise, structured, and practical."
            ),
            messages=[{"role": "user", "content": prompt}],
        )
        return resp.content[0].text
    return "AI not configured. Please add API keys to st.secrets."


# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
for key, default in {
    "df": None,
    "completed_map": {},
    "venture_meta": {},
    "chat_history": [],
    "selected_mentor": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ─────────────────────────────────────────────
# LOAD DATA WHEN FILE UPLOADED
# ─────────────────────────────────────────────
if uploaded_file:
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    with st.spinner("Reading tracker file..."):
        df, completed_map, venture_meta = load_tracker(file_bytes)
    st.session_state.df = df
    st.session_state.completed_map = completed_map
    st.session_state.venture_meta = venture_meta
    if not df.empty:
        st.sidebar.success(f"✅ Loaded {len(df)} feedback entries across {df['mentor'].nunique()} mentors.")
    else:
        st.sidebar.error("⚠️ No feedback data found. Check sheet names.")

df            = st.session_state.df
completed_map = st.session_state.completed_map

# ─────────────────────────────────────────────
# EMPTY STATE
# ─────────────────────────────────────────────
if df is None or (isinstance(df, pd.DataFrame) and df.empty):
    st.info("👈 Upload the Accelerate Mentor Connect Tracker Excel file from the sidebar to get started.")
    st.stop()

# ─────────────────────────────────────────────
# DERIVED DATA
# ─────────────────────────────────────────────
all_mentors  = sorted(df["mentor"].unique().tolist())
all_hubs     = sorted([h for h in df["hub"].unique()    if h])
all_programs = sorted([p for p in df["program"].unique() if p])

# ─────────────────────────────────────────────
# HELPER: RATING BADGE HTML
# ─────────────────────────────────────────────
RATING_STYLE = {
    "Good":     ("🟢", "#EAF3DE", "#27500A"),
    "Moderate": ("🟡", "#FAEEDA", "#633806"),
    "Poor":     ("🔴", "#FCEBEB", "#791F1F"),
}
RN_STYLE = {
    True:  ("By RN",     "#E6F1FB", "#0C447C"),
    False: ("Not by RN", "#EEEDFE", "#3C3489"),
}

def badge(label, bg, color):
    return (
        f"<span style='background:{bg};color:{color};padding:2px 9px;"
        f"border-radius:8px;font-size:12px;font-weight:600;'>{label}</span>"
    )

# ─────────────────────────────────────────────
# TOP SUMMARY METRICS
# ─────────────────────────────────────────────
st.markdown("---")
total_mentors  = df["mentor"].nunique()
total_good     = (df["rating"] == "Good").sum()
total_moderate = (df["rating"] == "Moderate").sum()
total_poor     = (df["rating"] == "Poor").sum()
total_not_rn   = (~df["by_rn"]).sum()

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("👤 Mentors",          total_mentors)
m2.metric("🟢 Good Feedbacks",   total_good)
m3.metric("🟡 Moderate",         total_moderate)
m4.metric("🔴 Poor Feedbacks",   total_poor)
m5.metric("🔵 Not by RN",        total_not_rn)
st.markdown("---")

# ─────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["🔍 Mentor Lookup", "📊 All Mentors View", "🤖 AI Insights"])

# ════════════════════════════════════════════
# TAB 1 — MENTOR LOOKUP (search + filters)
# ════════════════════════════════════════════
with tab1:
    st.subheader("Search & Filter Mentor Feedback")

    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
    with c1:
        mentor_search = st.text_input("🔎 Search mentor or venture name", placeholder="e.g. Sriram or Bebe Burp")
    with c2:
        rating_filter = st.multiselect("Rating", ["Good", "Moderate", "Poor"])
    with c3:
        hub_filter = st.multiselect("Hub", all_hubs)
    with c4:
        program_filter = st.multiselect("Program", all_programs)

    # Apply filters
    filtered = df.copy()
    if mentor_search:
        q = mentor_search.lower()
        filtered = filtered[
            filtered["mentor"].str.lower().str.contains(q, na=False) |
            filtered["venture"].str.lower().str.contains(q, na=False)
        ]
    if rating_filter:
        filtered = filtered[filtered["rating"].isin(rating_filter)]
    if hub_filter:
        filtered = filtered[filtered["hub"].isin(hub_filter)]
    if program_filter:
        filtered = filtered[filtered["program"].isin(program_filter)]

    mentors_in_view = sorted(filtered["mentor"].unique().tolist())
    st.caption(f"Showing **{len(mentors_in_view)} mentor(s)** · **{len(filtered)} feedback entries**")

    if not mentors_in_view:
        st.warning("No mentors match the current filters.")
        st.stop()

    # ── MENTOR CARDS ─────────────────────────────────────────────────
    for mentor_name in mentors_in_view:
        mentor_df = filtered[filtered["mentor"] == mentor_name]
        all_mentor_df = df[df["mentor"] == mentor_name]  # unfiltered for counts

        g = (all_mentor_df["rating"] == "Good").sum()
        m = (all_mentor_df["rating"] == "Moderate").sum()
        p = (all_mentor_df["rating"] == "Poor").sum()
        completed = completed_map.get(mentor_name, 0)
        total_fb  = len(all_mentor_df)

        # Card header colour: mostly good → green tint, any poor → red tint, else neutral
        if p > 0:
            header_bg = "#FFF0F0"
        elif g >= m:
            header_bg = "#F0F7EB"
        else:
            header_bg = "#FFFBF0"

        with st.expander(
            f"**{mentor_name}**  ·  🟢 {g}  🟡 {m}  🔴 {p}  |  "
            f"📅 {completed} meetings completed  ·  💬 {total_fb} feedbacks",
            expanded=False,
        ):
            # ── Summary pill row ────────────────────────────────────
            pills_html = (
                f"&nbsp;"
                + badge(f"🟢 {g} Good",     "#EAF3DE", "#27500A")
                + "&nbsp;"
                + badge(f"🟡 {m} Moderate",  "#FAEEDA", "#633806")
                + "&nbsp;"
                + badge(f"🔴 {p} Poor",      "#FCEBEB", "#791F1F")
                + "&nbsp;&nbsp;"
                + badge(f"📅 {completed} meetings", "#E6F1FB", "#0C447C")
            )
            st.markdown(pills_html, unsafe_allow_html=True)
            st.markdown("")

            # ── Venture entries (from filtered set) ──────────────────
            for _, row in mentor_df.sort_values("rating", key=lambda s: s.map({"Poor":0,"Moderate":1,"Good":2})).iterrows():
                r_icon, r_bg, r_color = RATING_STYLE[row["rating"]]
                rn_label, rn_bg, rn_color = RN_STYLE[row["by_rn"]]

                # Left accent border via container trick using st.container + markdown
                st.markdown(
                    f"""<div style='border-left:3px solid {r_color};
                        background:{r_bg}20;
                        border-radius:6px;
                        padding:10px 14px;
                        margin-bottom:8px;'>
                        <div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:4px;'>
                            <span style='font-weight:600;font-size:14px;color:#1F3864;'>{row['venture']}</span>
                            {badge(row['rating'], r_bg, r_color)}
                            {badge(rn_label, rn_bg, rn_color)}
                            {"" if not row['meet_again'] else badge("Re-meet: " + row['meet_again'], "#F3E5F5", "#3C3489")}
                        </div>
                        {"<p style='font-size:13px;color:#444;font-style:italic;margin:2px 0 0 0;'>\"" + row['feedback'] + "\"</p>" if row['feedback'] else "<p style='font-size:12px;color:#999;margin:2px 0 0 0;'>No founder comment.</p>"}
                        {"<p style='font-size:12px;color:#555;margin:4px 0 0 0;'><b>RN Remarks:</b> " + row['rn_remarks'] + "</p>" if row['rn_remarks'] else ""}
                        {"<p style='font-size:12px;color:#555;margin:2px 0 0 0;'><b>Action Items:</b> " + row['action_items'] + "</p>" if row['action_items'] else ""}
                        {"<p style='font-size:11px;color:#888;margin:4px 0 0 0;'>Hub: " + row['hub'] + " · Program: " + row['program'] + "</p>" if row['hub'] or row['program'] else ""}
                    </div>""",
                    unsafe_allow_html=True,
                )

            # ── AI Quick Insight button ──────────────────────────────
            if st.button(f"🤖 Get AI insight for {mentor_name}", key=f"ai_{mentor_name}"):
                with st.spinner("Generating insight..."):
                    entries_text = "\n".join([
                        f"- Venture: {row['venture']} | Rating: {row['rating']} | "
                        f"By RN: {'Yes' if row['by_rn'] else 'No'} | "
                        f"Feedback: {row['feedback'] or 'None'} | "
                        f"RN Remarks: {row['rn_remarks'] or 'None'}"
                        for _, row in all_mentor_df.iterrows()
                    ])
                    prompt = f"""Mentor: {mentor_name}
Completed meetings: {completed}
Total feedbacks: {total_fb} (Good: {g}, Moderate: {m}, Poor: {p})

Feedback entries:
{entries_text}

Provide a concise 3-part analysis:
1. Overall pattern (2-3 lines)
2. Strengths observed (bullet points)
3. Concerns / action items for the RN team (bullet points)

Be specific and use the actual feedback content. Keep it practical for program ops."""
                    insight = call_ai(prompt)
                    st.markdown("**🤖 AI Insight:**")
                    st.markdown(insight)


# ════════════════════════════════════════════
# TAB 2 — ALL MENTORS TABLE VIEW
# ════════════════════════════════════════════
with tab2:
    st.subheader("All Mentors — Summary Table")

    # Build summary per mentor
    summary_rows = []
    for mentor_name in all_mentors:
        mdf = df[df["mentor"] == mentor_name]
        g = (mdf["rating"] == "Good").sum()
        m = (mdf["rating"] == "Moderate").sum()
        p = (mdf["rating"] == "Poor").sum()
        not_rn = (~mdf["by_rn"]).sum()
        completed = completed_map.get(mentor_name, 0)
        ventures  = ", ".join(sorted(mdf["venture"].unique()))
        hubs_list = ", ".join(sorted([h for h in mdf["hub"].unique() if h]))
        progs_list= ", ".join(sorted([pr for pr in mdf["program"].unique() if pr]))

        summary_rows.append({
            "Mentor":           mentor_name,
            "Meetings Done":    completed,
            "Total Feedbacks":  len(mdf),
            "🟢 Good":          g,
            "🟡 Moderate":      m,
            "🔴 Poor":          p,
            "Not by RN":        not_rn,
            "Ventures":         ventures,
            "Hub(s)":           hubs_list,
            "Program(s)":       progs_list,
        })

    summary_df = pd.DataFrame(summary_rows)

    # Filters
    fc1, fc2, fc3 = st.columns([1,1,1])
    with fc1:
        sort_col = st.selectbox("Sort by", ["🟢 Good","Meetings Done","Total Feedbacks","🔴 Poor","🟡 Moderate"], index=0)
    with fc2:
        sort_dir = st.radio("Order", ["Descending","Ascending"], horizontal=True)
    with fc3:
        hub_f2 = st.multiselect("Filter by Hub", all_hubs, key="hub_tab2")

    disp = summary_df.copy()
    if hub_f2:
        disp = disp[disp["Hub(s)"].apply(lambda x: any(h in x for h in hub_f2))]
    disp = disp.sort_values(sort_col, ascending=(sort_dir == "Ascending"))

    st.dataframe(disp, use_container_width=True, hide_index=True)

    # Download button
    csv = disp.to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Download as CSV",
        data=csv,
        file_name="mentor_feedback_summary.csv",
        mime="text/csv",
    )


# ════════════════════════════════════════════
# TAB 3 — AI INSIGHTS CHAT
# ════════════════════════════════════════════
with tab3:
    st.subheader("🤖 AI Insights — Ask anything about the feedback data")
    st.caption("Ask questions like: *Which mentors have poor feedback from ventures we connected?* or *Summarise Chennai hub feedback.*")

    # Render chat history
    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    user_q = st.chat_input("Ask about mentor feedback, patterns, or specific ventures...")

    if user_q:
        with st.chat_message("user"):
            st.markdown(user_q)
        st.session_state.chat_history.append({"role": "user", "content": user_q})

        with st.chat_message("assistant"):
            with st.spinner("Analysing..."):

                # Build compact data context for AI
                mentor_summaries = []
                for mentor_name in all_mentors:
                    mdf = df[df["mentor"] == mentor_name]
                    g = (mdf["rating"] == "Good").sum()
                    m = (mdf["rating"] == "Moderate").sum()
                    p = (mdf["rating"] == "Poor").sum()
                    completed = completed_map.get(mentor_name, 0)
                    entries = []
                    for _, row in mdf.iterrows():
                        entries.append(
                            f"    • {row['venture']} [{row['rating']}]"
                            f"[{'By RN' if row['by_rn'] else 'Not by RN'}]"
                            f"[Hub:{row['hub']}][Prog:{row['program']}]"
                            f": {row['feedback'][:120] if row['feedback'] else 'no comment'}"
                            + (f" | RN: {row['rn_remarks'][:100]}" if row['rn_remarks'] else "")
                        )
                    mentor_summaries.append(
                        f"MENTOR: {mentor_name} | Done:{completed} | G:{g} M:{m} P:{p}\n"
                        + "\n".join(entries)
                    )

                data_context = "\n\n".join(mentor_summaries[:40])  # cap to avoid token overflow

                # Previous chat context
                prev_chat = ""
                for msg in st.session_state.chat_history[-6:]:
                    role = "User" if msg["role"] == "user" else "Assistant"
                    prev_chat += f"{role}: {msg['content']}\n"

                prompt = f"""You are assisting the NEN Resources Network team with mentor feedback analysis.

DATA SNAPSHOT (Feedback from Founders + Mentor Connects):
{data_context}

PREVIOUS CONVERSATION:
{prev_chat}

USER QUESTION: {user_q}

Instructions:
- Answer specifically using the data provided
- Use mentor names, venture names, ratings and feedback text in your answer
- Be structured: use bullet points or sections where helpful
- Flag any concerns (poor feedback, not-by-RN sessions, re-meet requests) proactively
- Keep response concise and actionable for program ops
"""
                response = call_ai(prompt, max_tokens=1500)
                st.markdown(response)
                st.session_state.chat_history.append({"role": "assistant", "content": response})
